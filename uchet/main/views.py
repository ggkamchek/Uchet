from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import UserPassesTestMixin
from django.utils.decorators import method_decorator
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.utils import timezone
from django.conf import settings
import os
import uuid
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.views.decorators.http import require_GET
import json
import math
from collections import defaultdict
from urllib.parse import quote, urlencode

from .models import User, Work, Criterion, Field, Level, Achievement, AchievementFieldValue, Period
from django.utils.safestring import mark_safe
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEACHER_DEPARTMENT_PRESETS = ('ДПИ', 'ИТВТ', 'СП')

def _department_choices_for_form():
    from_db = list(
        User.objects.exclude(department='')
        .values_list('department', flat=True)
        .distinct()
    )
    merged = sorted(set(TEACHER_DEPARTMENT_PRESETS) | set(from_db))
    return merged

def _department_display_label(user):
    text = (getattr(user, 'department', None) or '').strip()
    return text if text else 'Без отделения'

def _director_achievement_queryset(request):
    period_id = request.GET.get('period')
    department = request.GET.get('department')
    work_id = request.GET.get('work')
    qs = Achievement.objects.select_related('user', 'period', 'work').prefetch_related(
        'field_values__field'
    )
    if period_id:
        qs = qs.filter(period_id=period_id)
    if department:
        qs = qs.filter(user__department=department)
    if work_id:
        qs = qs.filter(work_id=work_id)
    return qs

def _user_fio_name_patronymic_last(user):
    parts = [user.first_name or '', user.patronymic or '', user.last_name or '']
    s = ' '.join(p for p in parts if p).strip()
    return s if s else user.username

def _sanitize_excel_sheet_title(name):
    for ch in '[]*\\/:?':
        name = name.replace(ch, ' ')
    name = (name or '').strip() or 'Преподаватель'
    return name[:31]

def _unique_excel_sheet_title(base, used_titles):
    title = _sanitize_excel_sheet_title(base)
    if title not in used_titles:
        used_titles.add(title)
        return title
    n = 2
    while True:
        suffix = f' ({n})'
        cand = _sanitize_excel_sheet_title(base[: 31 - len(suffix)] + suffix)
        if cand not in used_titles:
            used_titles.add(cand)
            return cand
        n += 1

def _level_caption_map_for_achievements(achievements):
    ids = []
    for ach in achievements:
        for fv in ach.field_values.all():
            if fv.field.type == Field.TypeChoices.CHOOSER:
                try:
                    ids.append(int(fv.value))
                except (TypeError, ValueError):
                    pass
    if not ids:
        return {}
    return {lvl.pk: lvl.caption for lvl in Level.objects.filter(pk__in=ids)}

def _export_field_value_display(field_value, level_captions):
    f = field_value.field
    raw = field_value.value or ''
    if f.type == Field.TypeChoices.CHOOSER:
        try:
            pk = int(raw)
            return level_captions.get(pk, raw)
        except (TypeError, ValueError):
            return raw
    if f.type == Field.TypeChoices.PHOTO:
        if raw:
            return os.path.basename(raw)
        return '—'
    return raw if raw else '—'

def _achievement_criterion_label_for_export(ach):
    names = {fv.field.criterion.name for fv in ach.field_values.all()}
    if names:
        return ', '.join(sorted(names))
    if ach.work:
        return ach.work.name
    return 'Критерий'

def _sanitize_filename_segment(text, max_len=80):
    text = (text or '').strip()
    for ch in '<>:"/\\|?*\r\n\t':
        text = text.replace(ch, '_')
    text = text.replace(' ', '_')
    while '__' in text:
        text = text.replace('__', '_')
    text = text.strip('_')
    if not text:
        text = 'без_названия'
    return text[:max_len]

def _director_export_download_filename(request):
    period_id = request.GET.get('period')
    if period_id:
        try:
            p = Period.objects.get(pk=period_id)
            period_part = _sanitize_filename_segment(p.name)
        except (Period.DoesNotExist, ValueError, TypeError):
            period_part = 'Период'
    else:
        period_part = 'Все_периоды'
    date_part = timezone.localtime(timezone.now()).strftime('%Y-%m-%d')
    return f'AllPrep_{period_part}_{date_part}.xlsx', date_part

def _content_disposition_attachment(download_name, ascii_fallback):
    enc = quote(download_name, safe='')
    return f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{enc}"

def _director_report_data(request, achievements):
    total_score_sum = sum(a.total_score() for a in achievements)

    budget_raw = request.GET.get('budget', '').strip()
    point_price = None
    budget_value = None
    if request.GET.get('reset') == '1':
        budget_raw = ''
    elif budget_raw:
        try:
            budget_value = float(budget_raw.replace(',', '.'))
            if total_score_sum > 0:
                point_price = budget_value / total_score_sum
        except ValueError:
            budget_raw = ''

    level_map = _level_caption_map_for_achievements(achievements)

    by_user = defaultdict(list)
    for ach in achievements:
        by_user[ach.user_id].append(ach)

    teacher_rows = []
    for uid in sorted(
        by_user.keys(),
        key=lambda i: (
            (by_user[i][0].user.last_name or ''),
            (by_user[i][0].user.first_name or ''),
            (by_user[i][0].user.patronymic or ''),
        ),
    ):
        u_ach = by_user[uid]
        user = u_ach[0].user
        balls = sum(a.total_score() for a in u_ach)
        prem = (balls * point_price) if point_price is not None else None
        teacher_rows.append((user, u_ach, balls, prem))

    itog_rows = []
    for idx, (user, _ach_list, balls, prem) in enumerate(teacher_rows, start=1):
        itog_rows.append(
            {
                'idx': idx,
                'fio': _user_fio_name_patronymic_last(user),
                'balls': balls,
                'prem': prem,
            }
        )

    used_sheet_names = {'ИТОГ'}
    sheets = []
    for user, ach_list, balls, prem in teacher_rows:
        base_title = (user.last_name or '').strip() or user.username
        stitle = _unique_excel_sheet_title(base_title, used_sheet_names)
        blocks = []
        for ach in sorted(
            ach_list,
            key=lambda a: (
                a.time_of_addition.timestamp() if a.time_of_addition else 0,
                a.id,
            ),
        ):
            score = ach.total_score()
            crit_label = _achievement_criterion_label_for_export(ach)
            fvs = sorted(
                ach.field_values.all(),
                key=lambda fv: (fv.field.criterion_id, fv.field_id),
            )
            field_rows = [
                {
                    'caption': fv.field.caption,
                    'value': _export_field_value_display(fv, level_map),
                }
                for fv in fvs
            ]
            blocks.append(
                {
                    'criterion_label': crit_label,
                    'score': score,
                    'fields': field_rows,
                }
            )
        sheets.append(
            {
                'user': user,
                'ach_list': ach_list,
                'sheet_title': stitle,
                'fio': _user_fio_name_patronymic_last(user),
                'total_balls': balls,
                'prem': prem,
                'blocks': blocks,
            }
        )

    return {
        'total_score_sum': total_score_sum,
        'budget_value': budget_value,
        'point_price': point_price,
        'itog_rows': itog_rows,
        'sheets': sheets,
    }

_AP_SIDE_THIN = Side(style='thin', color='FF000000')
_AP_BORDER_THIN = Border(
    left=_AP_SIDE_THIN,
    right=_AP_SIDE_THIN,
    top=_AP_SIDE_THIN,
    bottom=_AP_SIDE_THIN,
)
_AP_SIDE_MED_F0 = Side(style='medium', color='FFF0F0F0')
_AP_SIDE_MED_DD = Side(style='medium', color='FFDDDDDD')
_AP_FILL_YELLOW = PatternFill('solid', fgColor='FFFFFF00')
_AP_FONT_CAL11 = Font(name='Calibri', size=11, bold=False, color='FF000000')
_AP_FONT_CAL11_RED = Font(name='Calibri', size=11, bold=False, color='FFFF0000')
_AP_FONT_CAL12 = Font(name='Calibri', size=12, bold=False, color='FF000000')
_AP_FONT_CAL12_B = Font(name='Calibri', size=12, bold=True, color='FF000000')
_AP_FONT_CAL14_B = Font(name='Calibri', size=14, bold=True, color='FF000000')
_AP_FONT_TNR135_B = Font(name='Times New Roman', size=13.5, bold=True, color='FF000000')
_AP_FONT_TNR14 = Font(name='Times New Roman', size=14, bold=False, color='FF000000')
_AP_FONT_TNR14_B = Font(name='Times New Roman', size=14, bold=True, color='FF000000')
_AP_FONT_TNR18_B = Font(name='Times New Roman', size=18, bold=True, color='FF000000')
_AP_ALIGN_CC = Alignment(horizontal='center', vertical='center', wrap_text=False)
_AP_ALIGN_CR = Alignment(horizontal='right', vertical='center', wrap_text=False)
_AP_ALIGN_LC_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
_AP_ALIGN_VCR = Alignment(horizontal='right', vertical='center', wrap_text=False)
_XLSX_NUM_BALL = '#,##0.000'
_XLSX_NUM_MONEY = '#,##0.00'
_XLSX_NUM_PRICE = '#,##0.0000'
_XLSX_NUM_BUDGET = '#,##0.00'
_AP_ITOG_COL_WIDTHS = {
    'A': 9.7109375,
    'B': 35.140625,
    'C': 24.5703125,
    'D': 25.85546875,
    'E': 24.85546875,
}
_AP_TEACHER_COL_A = 123.140625
_AP_TEACHER_COL_B = 14.0
_AP_BORDER_EMPTY = Border()

def _xlsx_row_height_for_wrapped_cell(value, col_width_units, min_h=15.0, max_h=280.0):
    if value is None:
        return min_h
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return min_h
    s = str(value).strip()
    if not s:
        return min_h
    eff_chars = max(5.0, col_width_units * 0.88)
    lines = max(1, math.ceil(len(s) / eff_chars))
    h = 12.0 + lines * 12.75
    return max(min_h, min(max_h, h))

def _xlsx_teacher_col_a_width_units(value, row_idx, max_w=200.0):
    if value is None:
        return _AP_TEACHER_COL_A
    s = str(value).strip()
    if not s:
        return _AP_TEACHER_COL_A
    if row_idx == 2:
        scale = 1.32
    elif row_idx == 4:
        scale = 1.18
    else:
        scale = 1.22
    units = 4.0
    for ch in s:
        units += (1.32 if ord(ch) > 127 else 1.18) * scale
    return min(max_w, max(_AP_TEACHER_COL_A, units))

def _style_director_xlsx_itog(ws, num_teacher_rows, total_row_idx):
    ws.freeze_panes = 'A3'
    for letter, w in _AP_ITOG_COL_WIDTHS.items():
        ws.column_dimensions[letter].width = w

    ws.row_dimensions[1].height = 48.75

    for col in range(2, 6):
        c = ws.cell(row=1, column=col)
        if c.value is None:
            continue
        c.font = _AP_FONT_CAL14_B
        if isinstance(c.value, str):
            c.alignment = _AP_ALIGN_CC
        elif col in (3, 5):
            c.fill = _AP_FILL_YELLOW
            c.alignment = _AP_ALIGN_VCR
            if col == 3:
                c.number_format = _XLSX_NUM_BUDGET
            else:
                c.number_format = _XLSX_NUM_PRICE
        else:
            c.alignment = _AP_ALIGN_CC

    for col in range(1, 5):
        c = ws.cell(row=2, column=col)
        c.font = _AP_FONT_CAL12_B
        c.border = _AP_BORDER_THIN
        c.alignment = _AP_ALIGN_CC

    last_data = 2 + max(num_teacher_rows, 0)
    for r in range(3, last_data + 1):
        ws.row_dimensions[r].height = 15.75
        for col in range(1, 5):
            c = ws.cell(row=r, column=col)
            c.font = _AP_FONT_CAL12
            c.border = _AP_BORDER_THIN
            if col == 1:
                c.alignment = _AP_ALIGN_CC
            elif col == 2:
                c.alignment = _AP_ALIGN_LC_WRAP
            else:
                c.alignment = _AP_ALIGN_CR
                c.number_format = _XLSX_NUM_BALL if col == 3 else _XLSX_NUM_MONEY

    if total_row_idx:
        ws.row_dimensions[total_row_idx].height = 15.75
        for col in range(1, 5):
            c = ws.cell(row=total_row_idx, column=col)
            c.font = _AP_FONT_CAL12
            c.border = _AP_BORDER_THIN
            if col == 1:
                c.alignment = _AP_ALIGN_CC
            elif col == 2:
                c.alignment = _AP_ALIGN_LC_WRAP
            elif col in (3, 4):
                c.fill = _AP_FILL_YELLOW
                c.alignment = _AP_ALIGN_CR
                c.number_format = _XLSX_NUM_BALL if col == 3 else _XLSX_NUM_MONEY

def _style_director_xlsx_teacher(ws, max_row):
    ws.column_dimensions['B'].width = _AP_TEACHER_COL_B
    if max_row >= 6:
        ws.freeze_panes = 'A6'

    ws['B1'].font = _AP_FONT_CAL11_RED
    ws['B1'].alignment = _AP_ALIGN_CR
    ws['B1'].number_format = _XLSX_NUM_BALL
    ws['B1'].border = _AP_BORDER_EMPTY

    ws['A2'].font = _AP_FONT_TNR18_B
    ws['A2'].alignment = _AP_ALIGN_LC_WRAP
    ws['A2'].border = _AP_BORDER_EMPTY
    ws.row_dimensions[2].height = 22.5

    ws['A4'].font = _AP_FONT_TNR135_B
    ws['A4'].alignment = _AP_ALIGN_LC_WRAP
    ws['A4'].border = _AP_BORDER_EMPTY
    ws.row_dimensions[4].height = 18.0

    for r in range(6, max_row + 1):
        ca = ws.cell(row=r, column=1)
        cb = ws.cell(row=r, column=2)
        ca.border = _AP_BORDER_EMPTY
        cb.border = _AP_BORDER_EMPTY

        if cb.value is not None and isinstance(cb.value, (int, float)):
            ca.font = _AP_FONT_TNR14_B
            ca.alignment = _AP_ALIGN_LC_WRAP
            cb.font = _AP_FONT_CAL11
            cb.alignment = _AP_ALIGN_CR
            cb.fill = _AP_FILL_YELLOW
            cb.number_format = _XLSX_NUM_BALL
            ws.row_dimensions[r].height = 19.5
        elif ca.value:
            ca.font = _AP_FONT_TNR14_B
            ca.alignment = _AP_ALIGN_LC_WRAP
            ca.border = Border(bottom=_AP_SIDE_MED_F0)
        else:
            ca.font = _AP_FONT_TNR14
            ca.alignment = _AP_ALIGN_LC_WRAP
            ca.border = Border(bottom=_AP_SIDE_MED_DD)
            ws.row_dimensions[r].height = 12.75

    w_a = _AP_TEACHER_COL_A
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip():
            w_a = max(w_a, _xlsx_teacher_col_a_width_units(v, r))
    ws.column_dimensions['A'].width = w_a

    if max_row >= 2 and ws['A2'].value:
        ws.row_dimensions[2].height = max(
            22.5,
            _xlsx_row_height_for_wrapped_cell(ws['A2'].value, w_a, min_h=22.5, max_h=90.0),
        )
    if max_row >= 4 and ws['A4'].value:
        ws.row_dimensions[4].height = max(
            18.0,
            _xlsx_row_height_for_wrapped_cell(ws['A4'].value, w_a, min_h=18.0, max_h=72.0),
        )

    for r in range(6, max_row + 1):
        ca = ws.cell(row=r, column=1)
        cb = ws.cell(row=r, column=2)
        is_ball = cb.value is not None and isinstance(cb.value, (int, float))
        if not is_ball and ca.value and str(ca.value).strip():
            ws.row_dimensions[r].height = _xlsx_row_height_for_wrapped_cell(
                ca.value, w_a, min_h=19.5, max_h=260.0
            )

def is_admin(user):
    return user.is_authenticated and user.role == User.RoleChoices.ADMIN

class AdminRequiredMixin:
    @method_decorator(user_passes_test(is_admin))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

def role_redirect(request):
    if request.user.is_authenticated:
        if request.user.role == User.RoleChoices.ADMIN:
            return redirect('dashboard_home')
        elif request.user.role == User.RoleChoices.USER:
            return redirect('teacher_dashboard')
        elif request.user.role == User.RoleChoices.DIRECTOR:
            return redirect('director_dashboard')
    return redirect('login')

@user_passes_test(is_admin)
def dashboard_home(request):
    context = {
        'latest_works': Work.objects.all().order_by('-id')[:5],
        'latest_criterions': Criterion.objects.all().order_by('-id')[:5],
        'latest_fields': Field.objects.all().order_by('-id')[:5],
        'latest_levels': Level.objects.all().order_by('-id')[:5],
        'latest_periods': Period.objects.all().order_by('-id')[:5],
    }
    return render(request, 'main/dashboard_home.html', context)

@user_passes_test(is_admin)
def user_create_page(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        role = request.POST.get('role', '')
        department = request.POST.get('department', '').strip()[:5]
        last_name = request.POST.get('last_name', '').strip()[:25]
        first_name = request.POST.get('first_name', '').strip()[:25]
        patronymic = request.POST.get('patronymic', '').strip()[:25]

        allowed_roles = {User.RoleChoices.USER, User.RoleChoices.DIRECTOR}
        if not username or not password or role not in allowed_roles:
            messages.error(request, 'Заполните логин, пароль и выберите корректную роль.')
            return redirect('user_create')

        if not last_name or not first_name or not patronymic:
            messages.error(request, 'Укажите фамилию, имя и отчество.')
            return redirect('user_create')

        if role == User.RoleChoices.USER and not department:
            messages.error(request, 'Для преподавателя укажите отделение.')
            return redirect('user_create')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Пользователь с таким логином уже существует.')
            return redirect('user_create')

        User.objects.create_user(
            username=username,
            password=password,
            role=role,
            first_name=first_name,
            last_name=last_name,
            patronymic=patronymic,
            department=department if role == User.RoleChoices.USER else (department or ''),
            email='',
        )
        messages.success(request, 'Пользователь успешно создан.')
        return redirect('user_create')

    users = User.objects.filter(role__in=[User.RoleChoices.USER, User.RoleChoices.DIRECTOR]).order_by('-id')
    return render(
        request,
        'main/user/user_create.html',
        {'users': users, 'department_choices': _department_choices_for_form()},
    )

class UserDeleteView(AdminRequiredMixin, DeleteView):
    model = User
    template_name = 'main/user/user_confirm_delete.html'
    success_url = reverse_lazy('user_create')

    def get_queryset(self):
        return User.objects.filter(
            role__in=[User.RoleChoices.USER, User.RoleChoices.DIRECTOR]
        )

    def dispatch(self, request, *args, **kwargs):
        if int(kwargs['pk']) == request.user.pk:
            messages.error(request, 'Нельзя удалить свою учётную запись.')
            return redirect('user_create')
        return super().dispatch(request, *args, **kwargs)

    def delete(self, request, *args, **kwargs):
        username = self.object.username
        messages.success(request, f'Пользователь «{username}» удалён.')
        return super().delete(request, *args, **kwargs)

def is_director(user):
    return user.is_authenticated and user.role == User.RoleChoices.DIRECTOR

@user_passes_test(is_director)
def director_dashboard(request):
    period_id = request.GET.get('period')
    department = request.GET.get('department')
    work_id = request.GET.get('work')

    achievements = list(_director_achievement_queryset(request))

    total_score = 0
    dept_totals = {}
    period_dept_totals = {}
    teacher_ids = set()

    for ach in achievements:
        score = ach.total_score()
        total_score += score
        dept = _department_display_label(ach.user)
        dept_totals[dept] = dept_totals.get(dept, 0) + score

        per_name = ach.period.name if ach.period else '—'
        period_dept_totals.setdefault(per_name, {})
        period_dept_totals[per_name][dept] = period_dept_totals[per_name].get(dept, 0) + score
        teacher_ids.add(ach.user_id)

    teachers_total = User.objects.filter(role=User.RoleChoices.USER).count() or 1
    teachers_count = len(teacher_ids)
    teachers_percent = round(teachers_count * 100 / teachers_total, 1)

    active_period = Period.objects.filter(status=True).first()

    budget_raw = request.GET.get('budget', '').strip()
    point_price = None
    if request.GET.get('reset') == '1':
        budget_raw = ''
    elif budget_raw:
        try:
            budget_value = float(budget_raw.replace(',', '.'))
            if total_score > 0:
                point_price = round(budget_value / total_score, 4)
        except ValueError:
            budget_raw = ''

    dept_labels = sorted(dept_totals.keys())
    dept_values = [round(dept_totals[d], 2) for d in dept_labels]

    period_labels = sorted(period_dept_totals.keys())
    dept_series = {}
    for per in period_labels:
        for dept in period_dept_totals[per].keys():
            dept_series.setdefault(dept, [])
    for dept in dept_series.keys():
        for per in period_labels:
            dept_series[dept].append(round(period_dept_totals.get(per, {}).get(dept, 0), 2))

    chart_data = {
        'bar': {
            'labels': dept_labels,
            'values': dept_values,
        },
        'line': {
            'labels': period_labels,
            'series': dept_series,
        },
    }

    periods = Period.objects.all().order_by('name')
    departments = sorted(
        User.objects.filter(role=User.RoleChoices.USER)
        .exclude(department='')
        .values_list('department', flat=True)
        .distinct()
    )
    works = Work.objects.all().order_by('name')

    report_period_label = 'Все периоды'
    if period_id:
        try:
            report_period_label = Period.objects.get(pk=int(period_id)).name
        except (Period.DoesNotExist, ValueError, TypeError):
            report_period_label = 'Все периоды'

    preview_sync_query = urlencode(
        [(k, request.GET.get(k, '')) for k in ('period', 'department', 'work', 'budget')]
    )

    rd = _director_report_data(request, achievements)
    total_prem = (
        rd['total_score_sum'] * rd['point_price']
        if rd['point_price'] is not None
        else None
    )
    report_preview = {
        'total_score_sum': rd['total_score_sum'],
        'budget_value': rd['budget_value'],
        'point_price': rd['point_price'],
        'total_prem': total_prem,
        'itog_rows': rd['itog_rows'],
        'teachers': [
            {
                'fio': s['fio'],
                'total_balls': s['total_balls'],
                'blocks': s['blocks'],
            }
            for s in rd['sheets']
        ],
    }

    context = {
        'director_total_score': f'{total_score:.3f}',
        'director_teachers_count': teachers_count,
        'director_teachers_percent': teachers_percent,
        'director_current_period_name': active_period.name if active_period else '',
        'director_report_period_label': report_period_label,
        'preview_sync_query': preview_sync_query,
        'chart_data_json': mark_safe(json.dumps(chart_data, ensure_ascii=False)),
        'periods': periods,
        'departments': departments,
        'works': works,
        'selected_period': period_id or '',
        'selected_department': department or '',
        'selected_work': work_id or '',
        'budget_value': budget_raw,
        'point_price': point_price,
        'report_preview': report_preview,
    }
    return render(request, 'main/director/director_dashboard.html', context)

@user_passes_test(is_director)
def director_export(request):
    achievements = list(_director_achievement_queryset(request))
    data = _director_report_data(request, achievements)

    total_score_sum = data['total_score_sum']
    budget_value = data['budget_value']
    point_price = data['point_price']
    itog_rows = data['itog_rows']
    sheets = data['sheets']

    wb = Workbook()
    ws_itog = wb.active
    ws_itog.title = 'ИТОГ'

    ws_itog['B1'] = 'Премиальный фонд'
    if budget_value is not None:
        ws_itog['C1'] = budget_value
    ws_itog['D1'] = 'Цена 1 балла'
    if point_price is not None:
        ws_itog['E1'] = point_price

    ws_itog['A2'] = '№'
    ws_itog['B2'] = 'ФИО'
    ws_itog['C2'] = 'Балл'
    ws_itog['D2'] = 'Премия'

    for row_data in itog_rows:
        r = row_data['idx'] + 2
        ws_itog.cell(row=r, column=1, value=row_data['idx'])
        ws_itog.cell(row=r, column=2, value=row_data['fio'])
        ws_itog.cell(row=r, column=3, value=row_data['balls'])
        if row_data['prem'] is not None:
            ws_itog.cell(row=r, column=4, value=row_data['prem'])

    total_row_idx = None
    if itog_rows:
        tr = 2 + len(itog_rows) + 1
        total_row_idx = tr
        ws_itog.cell(row=tr, column=3, value=total_score_sum)
        if point_price is not None:
            ws_itog.cell(row=tr, column=4, value=total_score_sum * point_price)

    _style_director_xlsx_itog(ws_itog, len(itog_rows), total_row_idx)

    level_map = _level_caption_map_for_achievements(achievements)
    for sheet in sheets:
        ach_list = sheet['ach_list']
        stitle = sheet['sheet_title']
        balls = sheet['total_balls']
        ws = wb.create_sheet(title=stitle)
        ws['B1'] = balls
        ws['A2'] = sheet['fio']
        ws['A4'] = 'Результаты работы'
        max_data_row = 4
        row = 6
        for ach in sorted(
            ach_list,
            key=lambda a: (
                a.time_of_addition.timestamp() if a.time_of_addition else 0,
                a.id,
            ),
        ):
            score = ach.total_score()
            crit_label = _achievement_criterion_label_for_export(ach)
            ws.cell(row=row, column=1, value=f'{crit_label} Баллы:\u00a0')
            ws.cell(row=row, column=2, value=score)
            max_data_row = row
            row += 1
            fvs = sorted(
                ach.field_values.all(),
                key=lambda fv: (fv.field.criterion_id, fv.field_id),
            )
            for fv in fvs:
                disp = _export_field_value_display(fv, level_map)
                ws.cell(row=row, column=1, value=f'{fv.field.caption}:{disp}')
                max_data_row = row
                row += 1
            row += 1
        _style_director_xlsx_teacher(ws, max_data_row)

    download_name, date_part = _director_export_download_filename(request)
    ascii_fallback = f'AllPrep_{date_part}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = _content_disposition_attachment(download_name, ascii_fallback)
    wb.save(response)
    return response

class WorkListView(AdminRequiredMixin, ListView):
    model = Work
    template_name = 'main/work/work_list.html'
    context_object_name = 'works'

class WorkCreateView(AdminRequiredMixin, CreateView):
    model = Work
    fields = ['name']
    template_name = 'main/work/work_form.html'
    success_url = reverse_lazy('work_list')

class WorkDetailView(AdminRequiredMixin, DetailView):
    model = Work
    template_name = 'main/work/work_detail.html'
    context_object_name = 'work'

class WorkUpdateView(AdminRequiredMixin, UpdateView):
    model = Work
    fields = ['name']
    template_name = 'main/work/work_form.html'
    success_url = reverse_lazy('work_list')

class WorkDeleteView(AdminRequiredMixin, DeleteView):
    model = Work
    template_name = 'main/work/work_confirm_delete.html'
    success_url = reverse_lazy('work_list')

class CriterionListView(AdminRequiredMixin, ListView):
    model = Criterion
    template_name = 'main/criterion/criterion_list.html'
    context_object_name = 'criterions'

class CriterionCreateView(AdminRequiredMixin, CreateView):
    model = Criterion
    fields = ['name', 'work', 'ratio']
    template_name = 'main/criterion/criterion_form.html'
    success_url = reverse_lazy('criterion_list')

    def get_initial(self):
        initial = super().get_initial()
        work_id = self.request.GET.get('work')
        if work_id:
            try:
                initial['work'] = Work.objects.get(pk=work_id)
            except Work.DoesNotExist:
                pass
        return initial

class CriterionDetailView(AdminRequiredMixin, DetailView):
    model = Criterion
    template_name = 'main/criterion/criterion_detail.html'
    context_object_name = 'criterion'

class CriterionUpdateView(AdminRequiredMixin, UpdateView):
    model = Criterion
    fields = ['name', 'work', 'ratio']
    template_name = 'main/criterion/criterion_form.html'
    success_url = reverse_lazy('criterion_list')

class CriterionDeleteView(AdminRequiredMixin, DeleteView):
    model = Criterion
    template_name = 'main/criterion/criterion_confirm_delete.html'
    success_url = reverse_lazy('criterion_list')

class FieldListView(AdminRequiredMixin, ListView):
    model = Field
    template_name = 'main/field/field_list.html'
    context_object_name = 'fields'

class FieldCreateView(AdminRequiredMixin, CreateView):
    model = Field
    fields = ['caption', 'criterion', 'type']
    template_name = 'main/field/field_form.html'
    success_url = reverse_lazy('field_list')

    def get_initial(self):
        initial = super().get_initial()
        criterion_id = self.request.GET.get('criterion')
        if criterion_id:
            try:
                initial['criterion'] = Criterion.objects.get(pk=criterion_id)
            except Criterion.DoesNotExist:
                pass
        return initial

class FieldDetailView(AdminRequiredMixin, DetailView):
    model = Field
    template_name = 'main/field/field_detail.html'
    context_object_name = 'field'

class FieldUpdateView(AdminRequiredMixin, UpdateView):
    model = Field
    fields = ['caption', 'criterion', 'type']
    template_name = 'main/field/field_form.html'
    success_url = reverse_lazy('field_list')

class FieldDeleteView(AdminRequiredMixin, DeleteView):
    model = Field
    template_name = 'main/field/field_confirm_delete.html'
    success_url = reverse_lazy('field_list')

class LevelListView(AdminRequiredMixin, ListView):
    model = Level
    template_name = 'main/level/level_list.html'
    context_object_name = 'levels'

class LevelCreateView(AdminRequiredMixin, CreateView):
    model = Level
    fields = ['caption', 'field', 'ratio']
    template_name = 'main/level/level_form.html'
    success_url = reverse_lazy('level_list')

    def get_initial(self):
        initial = super().get_initial()
        field_id = self.request.GET.get('field')
        if field_id:
            try:
                initial['field'] = Field.objects.get(pk=field_id)
            except Field.DoesNotExist:
                pass
        return initial

class LevelDetailView(AdminRequiredMixin, DetailView):
    model = Level
    template_name = 'main/level/level_detail.html'
    context_object_name = 'level'

class LevelUpdateView(AdminRequiredMixin, UpdateView):
    model = Level
    fields = ['caption', 'field', 'ratio']
    template_name = 'main/level/level_form.html'
    success_url = reverse_lazy('level_list')

class LevelDeleteView(AdminRequiredMixin, DeleteView):
    model = Level
    template_name = 'main/level/level_confirm_delete.html'
    success_url = reverse_lazy('level_list')

class PeriodListView(AdminRequiredMixin, ListView):
    model = Period
    template_name = 'main/period/period_list.html'
    context_object_name = 'periods'

class PeriodCreateView(AdminRequiredMixin, CreateView):
    model = Period
    fields = ['name', 'start_date', 'end_date', 'status']
    template_name = 'main/period/period_form.html'
    success_url = reverse_lazy('period_list')

class PeriodDetailView(AdminRequiredMixin, DetailView):
    model = Period
    template_name = 'main/period/period_detail.html'
    context_object_name = 'period'

class PeriodUpdateView(AdminRequiredMixin, UpdateView):
    model = Period
    fields = ['name', 'start_date', 'end_date', 'status']
    template_name = 'main/period/period_form.html'
    success_url = reverse_lazy('period_list')

class PeriodDeleteView(AdminRequiredMixin, DeleteView):
    model = Period
    template_name = 'main/period/period_confirm_delete.html'
    success_url = reverse_lazy('period_list')

class TeacherRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.role == User.RoleChoices.USER

def teacher_required(view_func):
    decorated_view_func = login_required(view_func)
    return user_passes_test(lambda u: u.is_authenticated and u.role == User.RoleChoices.USER)(decorated_view_func)

@teacher_required
def teacher_dashboard(request):

    storage = messages.get_messages(request)
    storage.used = True

    achievements = Achievement.objects.filter(user=request.user).select_related(
        'work', 'period'
    ).prefetch_related(
        'field_values__field',
        'field_values__field__levels'
    ).order_by('-time_of_addition')
    return render(request, 'main/teacher/teacher_dashboard.html', {'achievements': achievements})

class TeacherAchievementDeleteView(TeacherRequiredMixin, DeleteView):
    model = Achievement
    template_name = 'main/teacher/teacher_achievement_confirm_delete.html'
    success_url = reverse_lazy('teacher_dashboard')

    def get_queryset(self):

        return Achievement.objects.filter(user=self.request.user)

@teacher_required
def teacher_achievement_create_page(request):
    works = Work.objects.all()
    return render(request, 'main/teacher/teacher_achievement_create.html', {'works': works})

@teacher_required
@require_GET
def get_criteria_json(request, work_id):
    try:
        work = Work.objects.get(pk=work_id)
        criteria = work.criterions.all().values('id', 'name')
        return JsonResponse(list(criteria), safe=False)
    except Work.DoesNotExist:
        return JsonResponse({'error': 'Work not found'}, status=404)

@teacher_required
@require_GET
def get_fields_html(request, criterion_id):
    try:
        criterion = Criterion.objects.get(pk=criterion_id)
        fields = criterion.fields.all().prefetch_related('levels')
        return render(request, 'main/partials/fields_form.html', {'fields': fields})
    except Criterion.DoesNotExist:
        return HttpResponseBadRequest('Criterion not found')

@teacher_required
def teacher_achievement_save(request):

    storage = messages.get_messages(request)
    storage.used = True

    if request.method == 'POST':
        work_id = request.POST.get('work')
        criterion_id = request.POST.get('criterion')

        if not work_id or not criterion_id:
            messages.error(request, 'Выберите работу и критерий')
            return redirect('teacher_achievement_create')

        work = get_object_or_404(Work, pk=work_id)
        criterion = get_object_or_404(Criterion, pk=criterion_id, work=work)

        active_period = Period.objects.filter(status=True).first()
        if not active_period:
            messages.error(request, 'Нет активного периода')
            return redirect('teacher_achievement_create')

        achievement = Achievement.objects.create(
            user=request.user,
            work=work,
            period=active_period,
            time_of_addition=timezone.now()
        )

        fields = criterion.fields.all()
        for field in fields:
            field_key = f'field_{field.id}'
            if field.type == Field.TypeChoices.TEXT:
                value = request.POST.get(field_key, '').strip()
                if value:
                    AchievementFieldValue.objects.create(achievement=achievement, field=field, value=value)
            elif field.type == Field.TypeChoices.CHOOSER:
                level_id = request.POST.get(field_key)
                if level_id and Level.objects.filter(pk=level_id, field=field).exists():
                    AchievementFieldValue.objects.create(achievement=achievement, field=field, value=str(level_id))
            elif field.type == Field.TypeChoices.PHOTO:
                if field_key in request.FILES:
                    uploaded_file = request.FILES[field_key]

                    upload_dir = os.path.join(settings.MEDIA_ROOT, 'achievement_photos')
                    os.makedirs(upload_dir, exist_ok=True)
                    ext = os.path.splitext(uploaded_file.name)[1]
                    filename = f"{achievement.id}_{field.id}_{uuid.uuid4().hex}{ext}"
                    file_path = os.path.join(upload_dir, filename)
                    with open(file_path, 'wb+') as destination:
                        for chunk in uploaded_file.chunks():
                            destination.write(chunk)
                    relative_path = os.path.join('achievement_photos', filename)
                    AchievementFieldValue.objects.create(achievement=achievement, field=field, value=relative_path)

        messages.success(request, 'Достижение успешно добавлено')
        return redirect('teacher_dashboard')
    else:
        return redirect('teacher_achievement_create')

@teacher_required
def teacher_achievement_edit(request, pk):
    achievement = get_object_or_404(
        Achievement.objects.prefetch_related('field_values__field__criterion', 'field_values__field__levels'),
        pk=pk,
        user=request.user
    )

    existing_values = {fv.field_id: fv for fv in achievement.field_values.all()}
    first_value = achievement.field_values.first()
    criterion = first_value.field.criterion if first_value else None

    if not criterion:
        messages.error(request, 'Невозможно определить критерий достижения для редактирования.')
        return redirect('teacher_dashboard')

    fields = Field.objects.filter(criterion=criterion).prefetch_related('levels')

    if request.method == 'POST':
        for field in fields:
            field_key = f'field_{field.id}'
            field_value_obj = existing_values.get(field.id)

            if field.type == Field.TypeChoices.TEXT:
                value = request.POST.get(field_key, '').strip()
                if field_value_obj:
                    field_value_obj.value = value
                    field_value_obj.save(update_fields=['value'])
                elif value:
                    AchievementFieldValue.objects.create(
                        achievement=achievement,
                        field=field,
                        value=value
                    )

            elif field.type == Field.TypeChoices.CHOOSER:
                level_id = request.POST.get(field_key)
                if level_id and Level.objects.filter(pk=level_id, field=field).exists():
                    if field_value_obj:
                        field_value_obj.value = str(level_id)
                        field_value_obj.save(update_fields=['value'])
                    else:
                        AchievementFieldValue.objects.create(
                            achievement=achievement,
                            field=field,
                            value=str(level_id)
                        )

            elif field.type == Field.TypeChoices.PHOTO and field_key in request.FILES:
                uploaded_file = request.FILES[field_key]
                upload_dir = os.path.join(settings.MEDIA_ROOT, 'achievement_photos')
                os.makedirs(upload_dir, exist_ok=True)
                ext = os.path.splitext(uploaded_file.name)[1]
                filename = f"{achievement.id}_{field.id}_{uuid.uuid4().hex}{ext}"
                file_path = os.path.join(upload_dir, filename)

                with open(file_path, 'wb+') as destination:
                    for chunk in uploaded_file.chunks():
                        destination.write(chunk)

                relative_path = os.path.join('achievement_photos', filename)
                if field_value_obj:
                    field_value_obj.value = relative_path
                    field_value_obj.save(update_fields=['value'])
                else:
                    AchievementFieldValue.objects.create(
                        achievement=achievement,
                        field=field,
                        value=relative_path
                    )

        messages.success(request, 'Достижение успешно обновлено.')
        return redirect('teacher_dashboard')

    fields_data = []
    for field in fields:
        current = existing_values.get(field.id)
        fields_data.append({
            'field': field,
            'current': current,
            'current_value': current.value if current else '',
        })

    context = {
        'achievement': achievement,
        'fields_data': fields_data,
        'criterion': criterion,
    }
    return render(request, 'main/teacher/teacher_achievement_edit.html', context)
